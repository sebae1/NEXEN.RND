from __future__ import annotations

import openpyxl as xl

from collections import defaultdict
from enum import StrEnum
from sqlalchemy import Column, Integer, Float, String, DateTime, \
    ForeignKey, select, delete, exists, literal_column, literal
from sqlalchemy.orm import relationship, load_only, aliased, selectinload, joinedload, object_session, mapped_column, Mapped

from db.database import Base, Session
from util import ExceptionWithMessage

CASCADE = "CASCADE"
DELETE = "DELETE"
RESTRICT = "RESTRICT"
SET_NULL = "SET NULL"
SET_DEFAULT = "SET DEFAULT"

MAXIMUM_DEPTH_OF_CATEGORY = 3

class Nexen(Base):
    __tablename__ = "nexen"

    pk = Column(Integer, primary_key=True, index=True)
    version = Column(String(10), nullable=False)
    created_at = Column(DateTime, nullable=False)

class EnumRND(StrEnum):
    RESEARCH = "Research"
    DEVELOP = "Develop"

class EnumOE(StrEnum):
    COMMON = "공통비"
    RE = "RE"
    OE = "OE"

class CostCtr(Base):
    __tablename__ = "cost_ctr"

    code       : Mapped[str] = mapped_column(String(20), primary_key=True, index=True)
    name       : Mapped[str] = mapped_column(String(20), nullable=False)
    rnd        : Mapped[str] = mapped_column(String(20), nullable=False, default=EnumRND.RESEARCH)
    oe         : Mapped[str] = mapped_column(String(20), nullable=False, default=EnumOE.COMMON)
    parent_code: Mapped[str|None] = mapped_column(String(20), ForeignKey("cost_ctr.code", onupdate=CASCADE, ondelete=CASCADE), nullable=True)

    parent = relationship(
        "CostCtr",
        back_populates="children",
        remote_side=code
    )
    children = relationship(
        "CostCtr",
        back_populates="parent",
        cascade="all, delete-orphan",
        single_parent=True
    )

    @classmethod
    def get(cls, code: str, eager: bool = True) -> CostCtr:
        with Session() as session:
            if eager:
                stmt = (
                    select(cls)
                    .options(
                        joinedload(cls.parent),
                        selectinload(cls.children)
                    )
                    .where(cls.code == code)
                )
                ctr = session.execute(stmt).scalar_one()
            else:
                ctr = session.get(cls, code)
        return ctr

    @classmethod
    def get_root_ctr(cls) -> CostCtr|None:
        """루트 ctr 반환 (기본값: 중앙연구소(CTO))"""
        stmt = select(cls).where(cls.parent_code.is_(None))
        with Session() as session:
            ctr = session.scalars(stmt).first()
        return ctr

    @classmethod
    def get_all(cls) -> dict[str, CostCtr]:
        """트리 구조를 탐색하여 상위부터 차례로 정렬된 CostCtr 아이템들을 반환

        Returns:
            { code (str): CostCtr }
        """
        anchor = (
            select(
                cls.code.label("code"),
                cls.name.label("name"),
                cls.rnd.label("rnd"),
                cls.oe.label("oe"),
                cls.parent_code.label("parent_code"),
                literal_column("0").label("level")
            )
            .where(cls.parent_code.is_(None))
        )
        cte = anchor.cte(name="cost_ctr_tree", recursive=True)
        alias_cte = aliased(cte)
        recursive = (
            select(
                cls.code,
                cls.name,
                cls.rnd,
                cls.oe,
                cls.parent_code,
                (alias_cte.c.level + 1).label("level")
            )
            .where(cls.parent_code == alias_cte.c.code)
        )
        cte = cte.union_all(recursive)
        with Session() as session:
            ctrs = (
                session.query(cls)
                .options(load_only(cls.code, cls.name, cls.rnd, cls.oe, cls.parent_code))
                .join(cte, cls.code == cte.c.code)
                .order_by(cte.c.level, cte.c.parent_code, cte.c.code)
                .all()
            )
        ret = {ctr.code: ctr for ctr in ctrs}
        return ret

    @classmethod
    def add(cls, code: str, name: str, rnd: EnumRND, oe: EnumOE, parent_code: str|None) -> CostCtr:
        is_exists = cls.has_code(code)
        if is_exists:
            raise ExceptionWithMessage(f"이미 존재하는 통화 코드 입니다: {code}")
        with Session() as session:
            ctr = cls(
                code=code,
                name=name,
                rnd=rnd.value,
                oe=oe.value,
                parent_code=parent_code
            )
            session.add(ctr)
            session.commit()
        return ctr

    @classmethod
    def delete(cls, code: str):
        with Session() as session:
            stmt = delete(cls).where(cls.code == code)
            session.execute(stmt)
            session.commit()

    @classmethod
    def has_code(cls, code: str) -> bool:
        with Session() as session:
            stmt = select(exists().where(cls.code == code))
            return bool(session.scalar(stmt))

    def update(self, code: str, name: str, rnd: EnumRND|str, oe: EnumOE|str, parent_code: str|None) -> CostCtr:
        with Session() as session:
            obj = session.get(CostCtr, self.code)
            obj.code = code
            obj.name = name
            obj.rnd = rnd.value if isinstance(rnd, EnumRND) else rnd
            obj.oe = oe.value if isinstance(oe, EnumOE) else oe
            obj.parent_code = parent_code
            session.commit()
        return obj

    def get_path(self) -> list[CostCtr,]:
        """최상위 노드에서부터 자기 자신까지의 노드 경로를 반환"""
        # leaf(self)에서 시작 (lvl=0)
        base = (
            select(
                CostCtr.code.label("code"),
                CostCtr.parent_code.label("parent_code"),
                literal(0).label("lvl"),
            )
            .where(CostCtr.code == self.code)
            .cte(name="ctr_lineage", recursive=True)
        )

        # parent로 올라가며 carry
        rec = select(
            CostCtr.code,
            CostCtr.parent_code,
            (base.c.lvl + 1).label("lvl"),
        ).join(base, CostCtr.code == base.c.parent_code)

        lineage = base.union_all(rec)

        with Session() as session:
            rows = session.execute(
                select(CostCtr).join(lineage, CostCtr.code == lineage.c.code).order_by(lineage.c.lvl.desc())
            ).scalars().all()
            return rows

    def get_descendant(self) -> list[CostCtr,]:
        """자기 자신을 포함하여 모든 자손 노드들을 반환"""
        # base: self부터 시작
        base = (
            select(
                CostCtr.code.label("code"),
                CostCtr.parent_code.label("parent_code"),
                literal(0).label("lvl"),
            )
            .where(CostCtr.code == self.code)
            .cte(name="ctr_tree", recursive=True)
        )

        # rec: children 따라 내려감
        rec = select(
            CostCtr.code,
            CostCtr.parent_code,
            (base.c.lvl + 1).label("lvl"),
        ).join(base, CostCtr.parent_code == base.c.code)

        tree = base.union_all(rec)

        # 정렬: lvl asc, code asc
        with Session() as session:
            rows = session.execute(
                select(CostCtr).join(tree, CostCtr.code == tree.c.code).order_by(tree.c.lvl, tree.c.code)
            ).scalars().all()
            return rows

class CostCategory(Base):
    __tablename__ = "cost_category"

    pk       : Mapped[int] = mapped_column(Integer, primary_key=True, index=True, autoincrement=True)
    name     : Mapped[str] = mapped_column(String(50), nullable=False)
    parent_pk: Mapped[int|None] = mapped_column(Integer, ForeignKey("cost_category.pk", onupdate=CASCADE, ondelete=CASCADE), nullable=True)

    parent = relationship(
        "CostCategory",
        remote_side=[pk],
        back_populates="children"
    )
    children = relationship(
        "CostCategory",
        back_populates="parent",
        cascade="all, delete-orphan"
    )
    elements = relationship(
        "CostElement",
        back_populates="category",
        lazy="selectin",
        passive_deletes=True
    )

    @classmethod
    def get(cls, pk: int, eager: bool = True) -> CostCategory:
        with Session() as session:
            if eager:
                stmt = (
                    select(cls)
                    .options(
                        joinedload(cls.parent),
                        selectinload(cls.children),
                        selectinload(cls.elements)
                    )
                    .where(cls.pk == pk)
                )
                ctr = session.execute(stmt).scalar_one()
            else:
                ctr = session.get(cls, pk)
        return ctr

    @classmethod
    def get_root_category(cls, eager: bool = True) -> CostCategory|None:
        """루트 category 반환 (기본값: 전체)"""
        if eager:
            stmt = (
                select(cls)
                .options(
                    joinedload(cls.parent),
                    selectinload(cls.children),
                    selectinload(cls.elements)
                )
                .where(cls.parent_pk.is_(None))
            )
        else:
            stmt = select(cls).where(cls.parent_pk.is_(None))
        with Session() as session:
            cat = session.execute(stmt).scalar_one()
        return cat

    @classmethod
    def get_all(cls) -> dict[int, CostCategory]:
        """트리 구조를 탐색하여 상위부터 차례로 정렬된 CostCategory 아이템들을 반환

        Returns:
            { pk (int): CostCategory }
        """
        anchor = (
            select(
                cls.pk.label("pk"),
                cls.name.label("name"),
                cls.parent_pk.label("parent_pk"),
                literal_column("0").label("level")
            )
            .where(cls.parent_pk.is_(None))
        )
        cte = anchor.cte(name="cost_category_tree", recursive=True)
        alias_cte = aliased(cte)
        recursive = (
            select(
                cls.pk,
                cls.name,
                cls.parent_pk,
                (alias_cte.c.level + 1).label("level")
            )
            .where(cls.parent_pk == alias_cte.c.pk)
        )
        cte = cte.union_all(recursive)
        with Session() as session:
            cats = (
                session.query(cls)
                .options(
                    load_only(cls.pk, cls.name, cls.parent_pk),
                    joinedload(cls.parent),
                    selectinload(cls.children),
                    selectinload(cls.elements)
                )
                .join(cte, cls.pk == cte.c.pk)
                .order_by(cte.c.level, cte.c.name, cte.c.parent_pk)
                .all()
            )
        ret = {cat.pk: cat for cat in cats}
        return ret

    @classmethod
    def get_direct_development_cost(cls) -> CostCategory|None:
        """이름이 '직접개발비'인 CostCategory 반환"""
        stmt = (
            select(cls)
            .options(
                joinedload(cls.parent),
                selectinload(cls.children),
                selectinload(cls.elements)
            )
            .where(cls.name == "직접개발비")
        )
        with Session() as session:
            try:
                cat = session.execute(stmt).scalar_one()
            except:
                return
            else:
                return cat

    @classmethod
    def add(cls, name: str, parent_pk: int) -> CostCategory:
        parent_cat = cls.get(parent_pk)
        if name in [cat.name for cat in parent_cat.children]:
            raise ExceptionWithMessage(f"카테고리 그룹에 동일한 이름이 존재합니다: {name}")
        with Session() as session:
            ctr = cls(
                name=name,
                parent_pk=parent_pk
            )
            session.add(ctr)
            session.commit()
        return ctr

    @classmethod
    def delete(cls, pk: int):
        with Session() as session:
            stmt = delete(cls).where(cls.pk == pk)
            session.execute(stmt)
            session.commit()

    def update(self, name: str, parent_pk: int) -> CostCategory:
        with Session() as session:
            session.expire_on_commit = False
            obj = session.get(CostCategory, self.pk)
            obj.name = name
            obj.parent_pk = parent_pk
            session.flush()
            stmt = (
                select(CostCategory)
                .options(
                    joinedload(CostCategory.parent),
                    selectinload(CostCategory.children),
                )
                .where(CostCategory.pk == obj.pk)
            )
            obj = session.execute(stmt).scalar_one()
            session.commit()
        return obj

    def get_path(self) -> list[CostCategory,]:
        """최상위 노드에서부터 자기 자신까지의 노드 경로를 반환"""
        # leaf(self)에서 시작 (lvl=0)
        base = (
            select(
                CostCategory.pk.label("pk"),
                CostCategory.parent_pk.label("parent_pk"),
                literal(0).label("lvl"),
            )
            .where(CostCategory.pk == self.pk)
            .cte(name="ctr_lineage", recursive=True)
        )

        # parent로 올라가며 carry
        rec = select(
            CostCategory.pk,
            CostCategory.parent_pk,
            (base.c.lvl + 1).label("lvl"),
        ).join(base, CostCategory.pk == base.c.parent_pk)

        lineage = base.union_all(rec)

        with Session() as session:
            rows = session.execute(
                select(CostCategory).join(lineage, CostCategory.pk == lineage.c.pk).order_by(lineage.c.lvl.desc())
            ).scalars().all()
            return rows

    def get_descendant(self) -> list[CostCategory,]:
        """자기 자신을 포함하여 모든 자손 노드들을 반환"""
        # base: self부터 시작
        base = (
            select(
                CostCategory.pk.label("pk"),
                CostCategory.parent_pk.label("parent_pk"),
                literal(0).label("lvl"),
            )
            .where(CostCategory.pk == self.pk)
            .cte(name="ctr_tree", recursive=True)
        )

        # rec: children 따라 내려감
        rec = select(
            CostCategory.pk,
            CostCategory.parent_pk,
            (base.c.lvl + 1).label("lvl"),
        ).join(base, CostCategory.parent_pk == base.c.pk)

        tree = base.union_all(rec)

        # 정렬: lvl asc, code asc
        with Session() as session:
            rows = session.execute(
                select(CostCategory).join(tree, CostCategory.pk == tree.c.pk).order_by(tree.c.lvl, tree.c.pk)
            ).scalars().all()
            return rows

class CostElement(Base):
    __tablename__ = "cost_element"
    __allow_unmapped__ = True

    code       : Mapped[str]  = mapped_column(String(20), primary_key=True, index=True)
    category_pk: Mapped[int|None] = mapped_column(Integer, ForeignKey("cost_category.pk", onupdate=CASCADE, ondelete=SET_NULL), nullable=True)
    description: Mapped[str]  = mapped_column(String(100), nullable=False, default="")

    category = relationship(
        "CostCategory",
        back_populates="elements",
        foreign_keys=[category_pk],
        lazy="joined"
    )

    _category_tree: list[str]|None = None

    @property
    def category_tree(self) -> list[str]|None:
        """preload 되거나 캐시가 존재한다면 category name tree 반환
        그렇지 않고 세션이 붙어 있으면 즉시 계산하여 반환
        그 외의 경우 None 반환
        """
        if self._category_tree is not None:
            return self._category_tree
        # 세션에 붙어 있으면 즉석 계산, detach면 에러로 안내
        sess = object_session(self)
        if sess is None:
            raise None
        self._category_tree = self._compute_category_trees(sess, self.category_pk)[self.category_pk]
        return self._category_tree

    @classmethod
    def _compute_category_trees(
            cls,
            session,
            category_pks: int|list[int]
        ) -> dict[int, list[str]]:
        """
        여러 leaf category_pk에 대해, 각 leaf가 속한 트리의 이름 경로(루트→리프)를
        한 번의 재귀 CTE로 계산해 반환합니다.

        Returns:
            { leaf_pk: ["root", "…", "leaf"] }
        """
        if category_pks is None:
            return defaultdict(list)
        if isinstance(category_pks, int):
            category_pks = [category_pks,]
        # 고유한 유효 pk만 추출
        leaf_ids = [int(pk) for pk in set(category_pks) if pk is not None]
        if not leaf_ids:
            return defaultdict(list)

        cc = CostCategory.__table__

        # base: leaf 카테고리들 (lvl=0, leaf_pk=자기자신)
        base = (
            select(
                cc.c.pk.label("pk"),
                cc.c.parent_pk.label("parent_pk"),
                cc.c.name.label("name"),
                cc.c.pk.label("leaf_pk"),
                literal(0).label("lvl"),
            )
            .where(cc.c.pk.in_(leaf_ids))
            .cte(name="cat_lineage", recursive=True)
        )

        # rec: parent로 한 단계씩 올라가며 leaf_pk를 carry
        rec = select(
            cc.c.pk,
            cc.c.parent_pk,
            cc.c.name,
            base.c.leaf_pk,
            (base.c.lvl + 1).label("lvl"),
        ).join(base, cc.c.pk == base.c.parent_pk)

        lineage = base.union_all(rec)

        # leaf_pk별로 lvl desc 정렬 → root→leaf 순으로 이름 리스트 만들기
        rows = session.execute(
            select(lineage.c.leaf_pk, lineage.c.name, lineage.c.lvl)
            .order_by(lineage.c.leaf_pk.asc(), lineage.c.lvl.desc())
        ).all()

        out: dict[int, list[str]] = defaultdict(list)
        for leaf_pk, name, _lvl in rows:
            out[leaf_pk].append(name)

        # 만약 존재하지 않는 leaf_pk가 있었다면 빈 리스트로 채워 줄 수도 있음(옵션)
        for pk in leaf_ids:
            out.setdefault(pk, [])

        return dict(out)

    @classmethod
    def get(cls, code: str, eager: bool = True) -> CostElement:
        with Session() as session:
            if eager:
                stmt = select(cls).options(joinedload(cls.category)).where(cls.code == code)
                elem = session.execute(stmt).scalar_one_or_none()
                elem._category_tree = cls._compute_category_trees(session, elem.category_pk)
            else:
                elem = session.get(cls, code)
        return elem

    @classmethod
    def get_all(cls) -> dict[str, CostElement]:
        """
        Returns:
            { code (str): CostElement }
        """
        with Session() as session:
            elements = (
                session.query(cls)
                .options(
                    load_only(cls.code, cls.category_pk, cls.description),
                )
                .order_by(cls.code)
                .all()
            )
            pk_map = cls._compute_category_trees(session, (e.category_pk for e in elements))

            # 각 elem에 캐시 주입
            for e in elements:
                e._category_tree = pk_map.get(e.category_pk, [])

        ret = {elem.code: elem for elem in elements}
        return ret

    @classmethod
    def get_involved_in_categories(cls, categories: list[CostCategory,]) -> list[CostElement,]:
        """카테고리 목록에 포함되는 Element 목록 반환"""
        if not categories:
            return []
        if isinstance(categories, CostCategory):
            categories = [categories,]

        cat_pks = [cat.pk for cat in categories]

        with Session() as session:
            stmt = (
                select(cls)
                .where(cls.category_pk.in_(cat_pks))
                .order_by(cls.code)
            )
            # lazy="joined"라서 category는 자동 eager-load됨
            return session.execute(stmt).scalars().all()

    @classmethod
    def add(cls, code: str, category_pk: int|None) -> CostElement:
        with Session() as session:
            elem = cls(
                code=code,
                category_pk=category_pk
            )
            session.add(elem)
            session.commit()
            elem._category_tree = __class__._compute_category_trees(session, category_pk)[category_pk]
        return elem

    @classmethod
    def delete(cls, code: str):
        with Session() as session:
            stmt = delete(cls).where(cls.code == code)
            session.execute(stmt)
            session.commit()

    @classmethod
    def has_code(cls, code: str) -> bool:
        with Session() as session:
            stmt = select(exists().where(cls.code == code))
            return bool(session.scalar(stmt))

    def update(self, code: str, category_pk: int|None, description: str|None = None) -> CostElement:
        with Session() as session:
            obj = session.get(CostElement, self.code)
            obj.code = code
            obj.category_pk = category_pk
            obj._category_tree = __class__._compute_category_trees(session, category_pk)[category_pk]
            print(obj._category_tree)
            obj.description = description or obj.description
            session.commit()
        return obj

class Currency(Base):
    __tablename__ = "currency"

    code: Mapped[str] = mapped_column(String(5), primary_key=True, index=True)
    unit: Mapped[int] = mapped_column(Integer, nullable=False)
    q1  : Mapped[float] = mapped_column(Float, nullable=False)
    q2  : Mapped[float] = mapped_column(Float, nullable=False)
    q3  : Mapped[float] = mapped_column(Float, nullable=False)
    q4  : Mapped[float] = mapped_column(Float, nullable=False)

    @classmethod
    def get(cls, code: str) -> Currency:
        with Session() as session:
            currency = session.get(cls, code)
        return currency

    @classmethod
    def get_all(cls) -> dict[str, Currency]:
        """
        Returns:
            { code (str): Currency }
        """
        with Session() as session:
            currencies = (
                session.query(cls)
                .options(load_only(cls.code, cls.unit, cls.q1, cls.q2, cls.q3, cls.q4))
                .order_by(cls.code)
                .all()
            )
        ret = {curr.code: curr for curr in currencies}
        return ret
    
    @classmethod
    def add(cls, code: str, unit: float, q1: float, q2: float, q3: float, q4: float) -> Currency:
        is_exists = cls.has_code(code)
        if is_exists:
            raise ExceptionWithMessage(f"이미 존재하는 통화 코드 입니다: {code}")
        with Session() as session:
            currency = cls(
                code=code,
                unit=unit,
                q1=q1,
                q2=q2,
                q3=q3,
                q4=q4
            )
            session.add(currency)
            session.commit()
        return currency

    @classmethod
    def delete(cls, code: str):
        with Session() as session:
            stmt = delete(cls).where(cls.code == code)
            session.execute(stmt)
            session.commit()

    @classmethod
    def has_code(cls, code: str) -> bool:
        with Session() as session:
            stmt = select(exists().where(cls.code == code))
            return bool(session.scalar(stmt))

    def update(self, code: str, unit: float, q1: float, q2: float, q3: float, q4: float) -> Currency:
        with Session() as session:
            obj = session.get(Currency, self.code)
            obj.code = code
            obj.unit = unit
            obj.q1 = q1
            obj.q2 = q2
            obj.q3 = q3
            obj.q4 = q4
            session.commit()
        return obj

    def get_currency_of_month(self, month: int) -> float:
        quarter = month//4+1
        if quarter == 1:
            return self.q1
        elif quarter == 2:
            return self.q2
        elif quarter == 3:
            return self.q3
        elif quarter == 4:
            return self.q4
        raise ValueError(f"Invalid month: {month}")


def read_ctr_excel(excel_file_path: str) -> list[CostCtr]:
    root_code = "K710000"
    ret: list[CostCtr] = []
    wb = xl.load_workbook(excel_file_path, data_only=True)
    ws = wb.worksheets[0]
    columns: dict[str, int] = {
        "CC": -1,
        "부문": -1,
        "팀명": -1,
        "연구/개발": -1,
        "OE/RE/평가/공통비": -1,
    }
    for i, col in enumerate(ws.iter_cols()):
        cell = col[1]
        key = str(cell.value).strip()
        if key in columns:
            columns[key] = i
        if -1 not in list(columns.values()):
            break
    for key, idx in columns.items():
        assert idx > 0, f"<{key}> 열을 찾을 수 없습니다."
    part_vs_bs_code: dict[str, str] = {} # "부문"에서 처음으로 확인되는 team을 BS로 취급
    code_set = set()
    for row in ws.iter_rows(3):
        code = row[columns["CC"]].value
        part = row[columns["부문"]].value
        team = row[columns["팀명"]].value
        rnd  = row[columns["연구/개발"]].value
        oe   = row[columns["OE/RE/평가/공통비"]].value
        if None in [code, part, team, rnd, oe]:
            continue
        code = str(code)
        part = str(part)
        team = str(team)
        rnd  = str(rnd )
        oe   = str(oe  )
        assert code not in code_set, f"Ctr Code가 중복됩니다: {code}"
        code_set.add(code)
        if rnd == "연구":
            rnd = EnumRND.RESEARCH
        elif rnd == "개발":
            rnd = EnumRND.DEVELOP
        else:
            continue
        if oe == "RE":
            oe = EnumOE.RE
        elif oe == "OE":
            oe = EnumOE.OE
        else:
            oe = EnumOE.COMMON
        parent_code = part_vs_bs_code.get(part)
        if parent_code is None:
            part_vs_bs_code[part] = code
        if parent_code is None \
            and code != root_code:
            parent_code = root_code
        ret.append(CostCtr(
            code       =code       ,
            name       =team       ,
            rnd        =rnd        ,
            oe         =oe         ,
            parent_code=parent_code,
        ))
    wb.close()
    return ret

def read_element_excel(excel_file_path: str) -> tuple[list[CostCategory], list[CostElement]]:
    cats: list[CostCategory] = []
    cat1_by_name: dict[str, CostCategory] = {}
    cat2_by_name: dict[str, CostCategory] = {}
    cat3_by_name: dict[str, CostCategory] = {}
    elems: list[CostElement] = []
    wb = xl.load_workbook(excel_file_path, data_only=True)
    ws = wb.worksheets[0]
    columns: dict[str, int] = {
        "계정코드": -1,
        "LV1": -1,
        "LV2": -1,
        "LV3": -1,
        "계정과목 개요": -1,
    }
    for i, col in enumerate(ws.iter_cols()):
        cell = col[3]
        key = str(cell.value).strip()
        if key in columns:
            columns[key] = i
        if -1 not in list(columns.values()):
            break
    for key, idx in columns.items():
        assert idx > 0, f"<{key}> 열을 찾을 수 없습니다."
    pk = 0
    root_cat = CostCategory(pk=pk, name="전체", parent_pk=None)
    cats.append(root_cat)
    pk += 1
    code_set = set()
    for row in ws.iter_rows(5):
        code = row[columns["계정코드"]].value
        lv1  = row[columns["LV1"]].value
        lv2  = row[columns["LV2"]].value
        lv3  = row[columns["LV3"]].value
        desc = row[columns["계정과목 개요"]].value
        if None in [code, lv1, lv2, lv3]:
            continue
        code = str(code)
        lv1  = str(lv1 )
        lv2  = str(lv2 )
        lv3  = str(lv3 )
        desc = str(desc or "").replace("\n", " ")
        assert code not in code_set, f"Element Code가 중복됩니다: {code}"
        code_set.add(code)
        if lv1 not in cat1_by_name:
            cat1 = CostCategory(pk=pk, name=lv1, parent_pk=root_cat.pk)
            cat1_by_name[cat1.name] = cat1
            cats.append(cat1)
            pk += 1
        if lv2 not in cat2_by_name:
            cat2 = CostCategory(pk=pk, name=lv2, parent_pk=cat1_by_name[lv1].pk)
            cat2_by_name[cat2.name] = cat2
            cats.append(cat2)
            pk += 1
        if lv3 not in cat3_by_name:
            cat3 = CostCategory(pk=pk, name=lv3, parent_pk=cat2_by_name[lv2].pk)
            cat3_by_name[cat3.name] = cat3
            cats.append(cat3)
            pk += 1
        elems.append(CostElement(
            code=code,
            category_pk=cat3_by_name[lv3].pk,
            description=desc
        ))
    wb.close()
    return cats, elems

