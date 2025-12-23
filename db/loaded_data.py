import re
import hashlib
import openpyxl as xl
import numpy as np
import pandas as pd
from .models import Currency, CostCategory, CostElement, CostCtr
from util import ExceptionWithMessage

REQUIRED_COLUMNS = ( # 엑셀 로우 데이터에 반드시 존재해야 하는 컬럼
    "Cost Center",
    "Cost Element",
    "Currency",
    "Cel Name",
    "Type",
    "1월", "2월", "3월", "4월", "5월", "6월",
    "7월", "8월", "9월", "10월", "11월", "12월",
)

DF_COLUMNS = ( # DF에 저장할 컬럼
    ("SHA256", "category"),
    ("Cost Center", "category"),
    ("Cost Element", "category"),
    ("Currency", "category"),
    ("대계정", "category"),
    ("계정항목", "category"),
    # 현지 통화 기준 금액
    # Plan: 계획, Actual: 집행
    ("RawPlan(1)" , "float32"), ("RawActual(1)" , "float32"),
    ("RawPlan(2)" , "float32"), ("RawActual(2)" , "float32"),
    ("RawPlan(3)" , "float32"), ("RawActual(3)" , "float32"),
    ("RawPlan(4)" , "float32"), ("RawActual(4)" , "float32"),
    ("RawPlan(5)" , "float32"), ("RawActual(5)" , "float32"),
    ("RawPlan(6)" , "float32"), ("RawActual(6)" , "float32"),
    ("RawPlan(7)" , "float32"), ("RawActual(7)" , "float32"),
    ("RawPlan(8)" , "float32"), ("RawActual(8)" , "float32"),
    ("RawPlan(9)" , "float32"), ("RawActual(9)" , "float32"),
    ("RawPlan(10)", "float32"), ("RawActual(10)", "float32"),
    ("RawPlan(11)", "float32"), ("RawActual(11)", "float32"),
    ("RawPlan(12)", "float32"), ("RawActual(12)", "float32"),
    # 원화로 환산된 금액
    ("ConvPlan(1)" , "float32"), ("ConvActual(1)" , "float32"),
    ("ConvPlan(2)" , "float32"), ("ConvActual(2)" , "float32"),
    ("ConvPlan(3)" , "float32"), ("ConvActual(3)" , "float32"),
    ("ConvPlan(4)" , "float32"), ("ConvActual(4)" , "float32"),
    ("ConvPlan(5)" , "float32"), ("ConvActual(5)" , "float32"),
    ("ConvPlan(6)" , "float32"), ("ConvActual(6)" , "float32"),
    ("ConvPlan(7)" , "float32"), ("ConvActual(7)" , "float32"),
    ("ConvPlan(8)" , "float32"), ("ConvActual(8)" , "float32"),
    ("ConvPlan(9)" , "float32"), ("ConvActual(9)" , "float32"),
    ("ConvPlan(10)", "float32"), ("ConvActual(10)", "float32"),
    ("ConvPlan(11)", "float32"), ("ConvActual(11)", "float32"),
    ("ConvPlan(12)", "float32"), ("ConvActual(12)", "float32"),
)

class LoadedData:
    df: pd.DataFrame = pd.DataFrame(
        {col: pd.Series(dtype=dt) for col, dt in DF_COLUMNS},
        index=pd.Index([], name="Key", dtype="string") # SHA256 + 행 번호로 조합한 고유 key
    )

    file_hash: dict = {} # { hash (str): filepath (str) }

    cached_cost_category: dict[int, CostCategory] = {}
    cached_cost_element: dict[str, CostElement] = {}
    cached_cost_ctr: dict[str, CostCtr] = {}
    cached_currency: dict[str, Currency] = {}

    @classmethod
    def load_raw_file(cls, filepaths: str | list[str,]):
        """엑셀로 된 raw data 파일을 읽고 DF에 concatenate

        Args:
            filepaths
                읽고자 하는 단일 raw data 파일의 경로 또는 다수 파일들의 경로 리스트
        """
        if isinstance(filepaths, str):
            filepaths = [filepaths,]
        sha256_vs_filepath = {}
        filepath_vs_sha256 = {}
        for filepath in filepaths:
            with open(filepath, "rb") as f:
                digest = hashlib.file_digest(f, "sha256")
                sha256 = digest.hexdigest()
            if sha256 in cls.file_hash:
                raise ExceptionWithMessage(f"이미 로드된 파일입니다.\n\n{cls.file_hash[sha256]}")
            sha256_vs_filepath[sha256] = filepath
            filepath_vs_sha256[filepath] = sha256
        
        data_frames = [cls.df,]
        for filepath in filepaths:
            sha256 = filepath_vs_sha256[filepath]
            df = cls._load_old_format(filepath, sha256)
            if df is None:
                df = cls._load_new_format(filepath, sha256)
            data_frames.append(df)

        cls.df = pd.concat(data_frames)
        cls.file_hash.update(sha256_vs_filepath)

    @classmethod
    def _load_old_format(cls, filepath: str, sha256: str) -> pd.DataFrame|None:
        """예전 포맷의 파일을 읽고 DataFrame으로 반환
        올바른 포맷이 아닌 경우 None 반환
        """
        fixed_columns = {
            "Cost Ctr": 3,
            "Cost Elem.": 8,
            "Name": 11,
            "Currency": 12
        }
        required_columns = (
            "Plan(1)", "Plan(2)", "Plan(3)", "Plan(4)", "Plan(5)", "Plan(6)", "Plan(7)", "Plan(8)", "Plan(9)", "Plan(10)", "Plan(11)", "Plan(12)",
            "Actual(1)", "Actual(2)", "Actual(3)", "Actual(4)", "Actual(5)", "Actual(6)", "Actual(7)", "Actual(8)", "Actual(9)", "Actual(10)", "Actual(11)", "Actual(12)"
        )
        currencies = LoadedData.cached_currency
        try:
            wb = xl.load_workbook(filepath)
            ws = wb["연구소"]
            for key, idx_col in fixed_columns.items():
                assert key == ws.cell(row=1, column=idx_col+1).value.strip() 
            column_vs_idx = {key: None for key in required_columns}
            columns = set(required_columns)
            for idx_col, col in enumerate(ws.iter_cols()):
                key = col[0].value.strip()
                if key in column_vs_idx:
                    column_vs_idx[key] = idx_col
                    columns.remove(key)
                if not columns:
                    break
            assert None not in column_vs_idx.values()
            lines = []
            for idx_row, row in enumerate(ws.iter_rows(2)):
                line = {key: np.nan for key, _ in DF_COLUMNS}
                line["SHA256"] = sha256
                line["Key"] = f"{sha256}-{idx_row:>04}"
                currency_code = row[fixed_columns["Currency"]].value.strip()
                for key, idx_col in fixed_columns.items():
                    value = str(row[idx_col].value).strip()
                    if key == "Name":
                        names = value.split("-")
                        if len(names) == 0:
                            cat1, cat2 = "", ""
                        elif len(names) == 1:
                            cat1, cat2 = names[0], names[0]
                        else:
                            cat1, cat2 = names[0], "-".join(names[1:])
                        line["대계정"] = cat1
                        line["계정항목"] = cat2
                    elif key == "Cost Ctr":
                        line["Cost Center"] = value
                    elif key == "Cost Elem.":
                        line["Cost Element"] = value
                    else:
                        line[key] = value
                cost_element = line["Cost Element"]
                if cost_element is None:
                    continue
                cost_element = str(cost_element).strip()
                if not cost_element \
                    or cost_element.startswith("6") \
                    or cost_element.startswith("9") \
                    or cost_element.startswith("1"):
                    continue
                for key, idx_col in column_vs_idx.items():
                    value = str(row[idx_col].value).strip()
                    try:
                        amount = int(value.replace(",", ""))
                        if key.startswith("Plan("):
                            month = int(key[5:-1])
                            line[f"RawPlan({month})"] = amount
                            line[f"ConvPlan({month})"] = amount * currencies[currency_code].get_currency_of_month(month)
                        elif key.startswith("Actual("):
                            month = int(key[7:-1])
                            line[f"RawActual({month})"] = amount
                            line[f"ConvActual({month})"] = amount * currencies[currency_code].get_currency_of_month(month)
                    except:
                        continue
                lines.append(line)
            df = pd.DataFrame(lines).set_index("Key")
            return df
        except:
            return
        finally:
            try:
                wb.close()
            except:
                pass

    @classmethod
    def _load_new_format(cls, filepath: str, sha256: str) -> pd.DataFrame:
        regex_yymm = re.compile(r"(\d{2})\.(\d{2})") # yy.mm 형식
        currencies = LoadedData.cached_currency
        wb = xl.load_workbook(filepath)
        ws = wb.active
        column_vs_idx = {col: None for col in REQUIRED_COLUMNS}
        columns = set(REQUIRED_COLUMNS)
        for i, col in enumerate(ws.iter_cols()):
            if col[0].value in REQUIRED_COLUMNS:
                column_vs_idx[col[0].value] = i
                columns.remove(col[0].value)
            else:
                match = regex_yymm.fullmatch(col[0].value)
                if match:
                    month = match.group(2).lstrip("0") + "월"
                    if month in REQUIRED_COLUMNS:
                        column_vs_idx[month] = i
                        columns.remove(month)
            if not columns:
                break
        for column, idx in column_vs_idx.items():
            if idx is None:
                raise ExceptionWithMessage(f"파일에서 '{column}' 열을 찾을 수 없습니다.\n\n{filepath}")
        lines = {} # { key (str): dict }
        for key_idx, row in enumerate(ws.iter_rows(2)):
            cost_element = row[column_vs_idx["Cost Element"]].value
            if cost_element is None:
                continue
            cost_element = str(cost_element).strip()
            if not cost_element \
                or cost_element.startswith("6") \
                or cost_element.startswith("9") \
                or cost_element.startswith("1"):
                continue
            cost_ctr = row[column_vs_idx["Cost Center"]].value
            if cost_ctr is None:
                continue
            cell_name = row[column_vs_idx["Cel Name"]].value
            if cell_name is None:
                continue

            data_key = f"{cost_element}_{cost_ctr}_{cell_name}"
            if data_key not in lines:
                line = {key: np.nan for key, _ in DF_COLUMNS}
                line["Key"] = f"{sha256}-{key_idx:>04}"
                line["SHA256"] = sha256
                line["Cost Center"] = cost_ctr
                line["Cost Element"] = cost_element
                line["Currency"] = row[column_vs_idx["Currency"]].value
    
                names = cell_name.split("-")
                if len(names) == 0:
                    cat1, cat2 = "", ""
                elif len(names) == 1:
                    cat1, cat2 = names[0], names[0]
                else:
                    cat1, cat2 = names[0], "-".join(names[1:])
                line["대계정"] = cat1
                line["계정항목"] = cat2
                lines[data_key] = line
            line = lines[data_key]

            # 계획인지 실적인지 판별
            type_of_line = row[column_vs_idx["Type"]].value
            if type_of_line == "Budget":
                key = "Plan"
            elif type_of_line == "Actual Sum":
                key = "Actual"
            else:
                continue

            for month in range(1, 13):
                try:
                    line[f"Raw{key}({month})"] = float(str(row[column_vs_idx[f"{month}월"]].value).replace(",", ""))
                except:
                    pass
            currency = line["Currency"]
            if currency in currencies:
                for month in range(1, 13):
                    try:
                        line[f"Conv{key}({month})"] = float(str(row[column_vs_idx[f"{month}월"]].value).replace(",", "")) * currencies[currency].get_currency_of_month(month) / currencies[currency].unit
                    except:
                        pass

        wb.close()
        df = pd.DataFrame(lines.values()).set_index("Key")
        return df

    @classmethod
    def reload(cls):
        """현재 파일을 다시 로드함"""
        filepaths = list(cls.file_hash.values())
        if not filepaths:
            return
        data_frames = []
        for filepath in filepaths:
            with open(filepath, "rb") as f:
                digest = hashlib.file_digest(f, "sha256")
                sha256 = digest.hexdigest()
            df = cls._load_old_format(filepath, sha256)
            if df is None:
                df = cls._load_new_format(filepath, sha256)
            data_frames.append(df)
        cls.df = pd.concat(data_frames)

    @classmethod
    def remove_raw_data(cls, file_hash: str):
        """DF에서 특정 파일로부터 로드된 데이터들을 삭제
        
        Args:
            file_hash
                삭제하고자 하는 데이터의 출처 파일의 SHA256 해시
        """
        df = cls.df
        df.drop(df[df["SHA256"]==file_hash].index, inplace=True)
        del cls.file_hash[file_hash]

    @classmethod
    def get_all_currencies(cls) -> set[str,]:
        """현재 로드된 DF에 포함된 통화 코드들 반환"""
        return set(cls.df["Currency"].unique().tolist())

    @classmethod
    def update_currency(cls):
        """DB에서 환율 정보를 불러와 현재 DF의 Conv 컬럼들을 업데이트"""
        df = cls.df
        currencies = cls.cached_currency
        for curr in currencies.values():
            mask = df["Currency"] == curr.code
            for month in range(1, 13):
                df.loc[mask, f"ConvPlan({month})"] = df.loc[mask, f"RawPlan({month})"] * curr.get_currency_of_month(month) / curr.unit
                df.loc[mask, f"ConvActual({month})"] = df.loc[mask, f"RawActual({month})"] * curr.get_currency_of_month(month) / curr.unit
        mask = ~df["Currency"].isin(currencies.keys())
        for month in range(1, 13):
            df.loc[mask, f"ConvPlan({month})"] = np.nan
            df.loc[mask, f"ConvActual({month})"] = np.nan

    @classmethod
    def get_level_of_ctr_from_cache(cls, ctr: CostCtr) -> int:
        level = 1
        while True:
            parent_code = ctr.parent_code
            if parent_code is None:
                return level
            ctr = cls.cached_cost_ctr[parent_code]
            level += 1
    
    @classmethod
    def get_level_of_category_from_cache(cls, category: CostCategory) -> int:
        level = 1
        while True:
            parent_pk = category.parent_pk
            if parent_pk is None:
                return level
            category = cls.cached_cost_category[parent_pk]
            level += 1

    @classmethod
    def get_category_path_from_cache(cls, category: CostCategory) -> str:
        paths = []
        while True:
            paths.append(category.name)
            if category.parent_pk is None:
                return " > ".join(paths[::-1][1:])
            category = cls.cached_cost_category[category.parent_pk]

    @classmethod
    def get_first_category(cls, category: CostCategory) -> CostCategory|None:
        """level==1 (전체) 인 경우 None 반환"""
        while True:
            level = cls.get_level_of_category_from_cache(category)
            if level == 1:
                return
            if level == 2:
                return category
            category = cls.cached_cost_category[category.parent_pk]

    @classmethod
    def get_bs(cls, ctr: CostCtr) -> CostCtr|None:
        """루트 CTR(중앙연구소)를 넘기면 None, 그 외의 경우 소속된 BS 반환
        BS를 넘기면 자기 자신이 반환됨
        """
        level = cls.get_level_of_ctr_from_cache(ctr)
        if level == 1:
            return
        if level == 2:
            return ctr
        return cls.cached_cost_ctr[ctr.parent_code]

    @classmethod
    def get_available_mask(cls) -> pd.Series:
        """캐시를 참고하여 '분류' 건에 대한 마스크 반환"""
        df = cls.df
        mask = (df["Cost Center"].isin(LoadedData.cached_cost_ctr)) \
            & (df["Cost Element"].isin(LoadedData.cached_cost_element)) \
            & (df["Currency"].isin(LoadedData.cached_currency))
        return mask

    @classmethod
    def get_filtered_df(cls) -> pd.DataFrame:
        """'미분류' 건들을 제외한 df 반환"""
        return cls.df.loc[cls.get_available_mask()]

    @classmethod
    def cache_all(cls):
        cls.cache_ctr()
        cls.cache_element()
        cls.cache_category()
        cls.cache_currency()

    @classmethod
    def cache_ctr(cls):
        cls.cached_cost_ctr = CostCtr.get_all()

    @classmethod
    def cache_element(cls):
        cls.cached_cost_element = CostElement.get_all()

    @classmethod
    def cache_category(cls):
        cls.cached_cost_category = CostCategory.get_all()

    @classmethod
    def cache_currency(cls):
        cls.cached_currency = Currency.get_all()
