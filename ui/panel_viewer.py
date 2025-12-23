import time
import matplotlib.pyplot as plt
import numpy as np
import openpyxl as xl
import wx
import wx.dataview as DV

from traceback import format_exc
from typing import Literal
from threading import Thread
from dataclasses import dataclass
from matplotlib.backends.backend_wxagg import FigureCanvasWxAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.axes import Axes
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, PatternFill
from wx.lib.scrolledpanel import ScrolledPanel

from util import simplify_won, COLORMAP, Config
from db import CostCategory, CostCtr, MAXIMUM_DEPTH_OF_CATEGORY, CostElement, LoadedData
from ui.component import TreeListCtrl, TreeListModelBase, TreeListNode, \
    FONT_COLOR_LOW_PORTION, FONT_COLOR_MID_PORTION, FONT_COLOR_HIGH_PORTION, FONT_COLOR_NEGATIVE_VALUE, \
    OPENAI_MARK_SVG, CLAUDE_MARK_SVG
from ui.component.ai_analysis import DialogAIResult, DialogModels
from ai import (
    _CostCategory, _CostElement, _CostCtr, _BudgetByCtr, _BudgetByElement,
    get_prompts_for_ai, analyze_by_claude, analyze_by_gpt
)

class _Config:
    UNIT = "자동"
    PERIOD = "전체"

@dataclass
class ItemCategory:
    category: CostCategory
    plan: float|None = None
    actual: float|None = None
    rem: float|None = None
    exe: float|None = None

@dataclass
class ItemCtr:
    ctr: CostCtr
    total: bool # 합계
    plan: float|None = None
    actual: float|None = None
    rem: float|None = None
    exe: float|None = None

class ModelCategory(TreeListModelBase):
    def __init__(self):
        TreeListModelBase.__init__(self, 5)
    
    def GetValue(self, item, col):
        if not item:
            return ""
        
        node_id = int(item.GetID())
        if node_id not in self.nodes:
            return ""
        node = self.nodes[node_id]
        item: ItemCategory = node.item
        category: CostCategory = item.category
        match col:
            case 0:
                return category.name
            case 1:
                return "" if item.plan is None else simplify_won(item.plan, _Config.UNIT)
            case 2:
                return "" if item.actual is None else simplify_won(item.actual, _Config.UNIT)
            case 3:
                return "" if item.rem is None else simplify_won(item.rem, _Config.UNIT)
            case 4:
                return "" if item.exe is None else f"{item.exe*100:0.1f}"
        return ""

    def GetAttr(self, item, col, attr):
        if not item:
            return False
        node_id = int(item.GetID())
        if node_id not in self.nodes:
            return ""
        node = self.nodes[node_id]
        item: ItemCategory = node.item
        match col:
            case 1: # 계획
                color = wx.Colour(0, 0, 0)
                try:
                    if item.plan < 0:
                        color = wx.Colour(FONT_COLOR_NEGATIVE_VALUE)
                except:
                    pass
                attr.SetColour(color)
                return True
            case 2: # 실적
                color = wx.Colour(0, 0, 0)
                try:
                    if item.exe < 0:
                        color = wx.Colour(FONT_COLOR_NEGATIVE_VALUE)
                except:
                    pass
                attr.SetColour(color)
                return True
            case 3: # 잔액
                if item.rem is None:
                    color = wx.Colour(0, 0, 0)
                elif item.rem >= 0:
                    color = wx.Colour(0, 0, 0)
                else:
                    color = wx.Colour(FONT_COLOR_NEGATIVE_VALUE)
                attr.SetColour(color)
                return True
            case 4: # 집행률
                if item.exe is None:
                    color = wx.Colour(0, 0, 0)
                elif item.exe >= 0.9:
                    attr.SetBold(True)
                    color = wx.Colour(FONT_COLOR_HIGH_PORTION)
                elif item.exe >= 0.5:
                    attr.SetBold(True)
                    color = wx.Colour(FONT_COLOR_MID_PORTION)
                else:
                    color = wx.Colour(FONT_COLOR_LOW_PORTION)
                attr.SetColour(color)
                return True
        return False

class ModelCtr(TreeListModelBase):
    def __init__(self):
        TreeListModelBase.__init__(self, 8)
    
    def GetValue(self, item, col):
        if not item:
            return ""
        
        node_id = int(item.GetID())
        if node_id not in self.nodes:
            return ""
        node = self.nodes[node_id]
        item: ItemCtr = node.item
        ctr: CostCtr = item.ctr
        match col:
            case 0:
                return ctr.name
            case 1:
                return "" if item.total else ctr.code
            case 2:
                return "" if item.total else ctr.rnd
            case 3:
                return "" if item.total else ctr.oe
            case 4:
                return "" if item.plan is None else simplify_won(item.plan, _Config.UNIT)
            case 5:
                return "" if item.actual is None else simplify_won(item.actual, _Config.UNIT)
            case 6:
                return "" if item.rem is None else simplify_won(item.rem, _Config.UNIT)
            case 7:
                return "" if item.exe is None else f"{item.exe*100:0.1f}"
        return ""

    def GetAttr(self, item, col, attr):
        if not item:
            return False
        node_id = int(item.GetID())
        if node_id not in self.nodes:
            return ""
        node = self.nodes[node_id]
        item: ItemCategory = node.item
        match col:
            case 4: # 계획
                color = wx.Colour(0, 0, 0)
                try:
                    if item.plan < 0:
                        color = wx.Colour(FONT_COLOR_NEGATIVE_VALUE)
                except:
                    pass
                attr.SetColour(color)
                return True
            case 5: # 실적
                color = wx.Colour(0, 0, 0)
                try:
                    if item.exe < 0:
                        color = wx.Colour(FONT_COLOR_NEGATIVE_VALUE)
                except:
                    pass
                attr.SetColour(color)
                return True
            case 6: # 잔액
                if item.rem is None:
                    color = wx.Colour(0, 0, 0)
                elif item.rem >= 0:
                    color = wx.Colour(0, 0, 0)
                else:
                    color = wx.Colour(FONT_COLOR_NEGATIVE_VALUE)
                attr.SetColour(color)
                return True
            case 7: # 집행률
                if item.exe is None:
                    color = wx.Colour(0, 0, 0)
                elif item.exe >= 0.9:
                    attr.SetBold(True)
                    color = wx.Colour(FONT_COLOR_HIGH_PORTION)
                elif item.exe >= 0.5:
                    attr.SetBold(True)
                    color = wx.Colour(FONT_COLOR_MID_PORTION)
                else:
                    color = wx.Colour(FONT_COLOR_LOW_PORTION)
                attr.SetColour(color)
                return True
        return False

class TreeCategory(TreeListCtrl):
    def __init__(self, parent: wx.Panel):
        TreeListCtrl.__init__(
            self,
            parent,
            ModelCategory(),
            {
                "Cost Category": 200,
                "계획": 120,
                "실적": 120,
                "잔액": 120,
                "집행률(%)": 70
            }
        )

class TreeCtr(TreeListCtrl):
    def __init__(self, parent: wx.Panel):
        TreeListCtrl.__init__(
            self,
            parent,
            ModelCtr(),
            {
                "Cost Ctr": 200,
                "코드": 80,
                "개발 비중": 100,
                "OE 비중": 100,
                "계획": 120,
                "실적": 120,
                "잔액": 120,
                "집행률(%)": 70
            }
        )

DPI = 80
TITLE_FONTSIZE = 12
BAR_WIDTH = 0.5
PIE_OFFSET = -0.5
PIE_LEGEND_BBOX_TO_ANCHOR = (0.8, 0.5)
LEGEND_KWARGS = {
    "frameon": False,
    "labelspacing": 1.5
}

def hide_axis(ax: Axes):
    for spine in ax.spines.values():
        spine.set_visible(False)
    ax.tick_params(left=False, bottom=False)
    ax.set_xticklabels([])
    ax.set_yticklabels([])

def draw_multiple_bar(ax: Axes, title: str = None, data: dict[str, float] = {}):
    ax.clear()
    ax.set_title(title or "", fontsize=TITLE_FONTSIZE)
    hide_axis(ax)
    ax.set_xlim(-1, 1.5)
    data = dict(sorted(data.items(), key=lambda item: -item[1]))
    if "합계" in data:
        keys = list(data)
        keys.remove("합계")
        keys.insert(0, "합계")
        data = {key: data[key] for key in keys}
    x_labels = list(data)
    if not x_labels:
        for i in range(8):
            ax.bar(i, 1, bottom=0, color="gray", width=BAR_WIDTH)
        ax.set_xlim(-1, 1.1*8)
        return
    values = list(data.values())
    y = np.zeros(len(x_labels))
    p = ax.bar(x_labels, values, bottom=y, width=BAR_WIDTH)
    ax.bar_label(
        p,
        label_type="edge",
        labels=[simplify_won(val) for val in values]
    )
    ax.set_xlim(-1, 1.1*len(x_labels))
    ax.set_xticks(range(len(x_labels)))
    ax.set_xticklabels(x_labels, rotation=45)

class DialogChart(wx.Dialog):
    # 상수 정의
    BAR_HEIGHT_PX = 60
    VERTICAL_PADDING_PX = 30
    
    # ----------------------------------------------------------------------
    # 초기화 및 기본 설정
    # ----------------------------------------------------------------------

    def __init__(self, parent: wx.Window, title: str, items: list['ItemCategory|ItemCtr',]):
        wx.Dialog.__init__(
            self,
            parent,
            title=title,
            style=wx.DEFAULT_DIALOG_STYLE | wx.RESIZE_BORDER
        )
        self.__items = items

        self.__setup_ui()
        self.__bind_events()
        self.draw()

    # ----------------------------------------------------------------------
    # UI 구성 (Layout)
    # ----------------------------------------------------------------------

    def __setup_ui(self):
        sz_horz = wx.BoxSizer(wx.HORIZONTAL)
        
        # 1. 좌측 설정 패널
        pn_left = self.__create_left_panel()
        
        # 2. 우측 차트 패널 (직접 정의한 Canvas 포함)
        pn_right = self.__create_right_panel()

        sz_horz.Add(pn_left, 0, wx.EXPAND)
        sz_horz.Add(pn_right, 1, wx.EXPAND)
        self.SetSizer(sz_horz)

        self.SetSize((900, 600))
        self.SetMinSize((600, 400))
        self.CenterOnParent()
        
    def __create_left_panel(self) -> wx.Panel:
        """좌측 설정 패널과 위젯을 생성하고 저장합니다."""
        pn_left = wx.Panel(self, style=wx.BORDER_RAISED)
        pn_left.SetMinSize((200, -1))
        
        sz_vert = wx.BoxSizer(wx.VERTICAL)
        
        # 라디오 버튼 (표시 모드)
        self.__rb_both = wx.RadioButton(pn_left, label="모두", style=wx.RB_GROUP)
        self.__rb_plan = wx.RadioButton(pn_left, label="계획")
        self.__rb_actual = wx.RadioButton(pn_left, label="실적")
        
        sz_vert.AddMany((
            (self.__rb_both, 0), ((-1, 5), 0),
            (self.__rb_plan, 0), ((-1, 5), 0),
            (self.__rb_actual, 0), ((-1, 15), 0),
        ))
        
        # 합계 체크박스
        self.__ck_sum = wx.CheckBox(pn_left, label="합계")
        self.__ck_sum.SetValue(True)
        sz_vert.Add(self.__ck_sum, 0, wx.BOTTOM, 5)
        
        # 항목별 체크박스
        self.__cks: dict[str, wx.CheckBox] = {}
        for item in self.__items:
            # Type Check가 필요한 경우 (ItemCategory/ItemCtr는 외부 정의 타입으로 가정)
            obj = item.category if isinstance(item, ItemCategory) else item.ctr 
            ck = wx.CheckBox(pn_left, label=obj.name)
            ck.SetValue(True)
            sz_vert.Add(ck, 0, wx.BOTTOM, 5)
            self.__cks[obj.name] = ck
            
        sz_left = wx.BoxSizer(wx.HORIZONTAL)
        sz_left.Add(sz_vert, 1, wx.EXPAND | wx.ALL, 10)
        pn_left.SetSizer(sz_left)
        
        return pn_left

    def __create_right_panel(self) -> wx.Panel:
        """Matplotlib Figure와 Canvas를 직접 생성하여 스크롤 윈도우에 배치합니다."""
        pn_right = wx.Panel(self)
        pn_right.SetBackgroundColour(wx.WHITE)

        # 스크롤 윈도우 생성
        self.__scroll = wx.ScrolledWindow(pn_right, style=wx.VSCROLL)
        self.__scroll.SetScrollRate(0, 10)
        self.__scroll.AlwaysShowScrollbars(False, True)

        # 1. Figure 생성 (배경 흰색)
        self.fig = Figure(facecolor='white', constrained_layout=True) # 초기 사이즈는 임의 설정
        
        # 2. Axes 추가
        self.ax = self.fig.add_subplot(111)
        
        # 3. Canvas 생성 (ScrollWindow를 부모로)
        self.__canvas = FigureCanvas(self.__scroll, -1, self.fig)

        # 스크롤 패널의 Sizer 설정
        s_scroll = wx.BoxSizer(wx.VERTICAL)
        s_scroll.Add(self.__canvas, 0, wx.EXPAND) # 0으로 설정하여 Sizer가 늘리지 않게 함 (직접 크기 지정 예정)
        self.__scroll.SetSizer(s_scroll)

        # 우측 패널 전체 Sizer
        sz_right = wx.BoxSizer(wx.HORIZONTAL)
        sz_right.Add(self.__scroll, 1, wx.EXPAND | wx.ALL, 15)
        pn_right.SetSizer(sz_right)
        
        return pn_right

    # ----------------------------------------------------------------------
    # 이벤트 바인딩
    # ----------------------------------------------------------------------

    def __bind_events(self):
        # 라디오 버튼 이벤트
        for rb in [self.__rb_both, self.__rb_plan, self.__rb_actual]:
            rb.Bind(wx.EVT_RADIOBUTTON, self.__on_control_changed)
            
        # 체크박스 이벤트 (합계 및 항목)
        self.__ck_sum.Bind(wx.EVT_CHECKBOX, self.__on_control_changed)
        for ck in self.__cks.values():
            ck.Bind(wx.EVT_CHECKBOX, self.__on_control_changed)

        self.__canvas.Bind(wx.EVT_RIGHT_DOWN, self.__on_right_click)
        self.Bind(wx.EVT_SIZE, self.__on_resize)
        self.Bind(wx.EVT_WINDOW_DESTROY, self.__on_destroy)
    
    def __on_control_changed(self, event):
        """라디오 버튼이나 체크박스 상태가 변경될 때 차트를 다시 그립니다."""
        self.draw()

    def __on_destroy(self, event):
        plt.close(self.fig)
        event.Skip()

    def __on_resize(self, event):
        if self.__data_to_draw:
            self.__update_canvas_size()
        event.Skip()

    def __on_right_click(self, event):
        menu = wx.Menu()
        save_item = menu.Append(wx.ID_ANY, "이미지 저장")
        self.Bind(wx.EVT_MENU, self.__on_save_chart, save_item)
        self.PopupMenu(menu)
        menu.Destroy()

    def __on_save_chart(self, event):
        dlg = wx.FileDialog(None, "이미지 저장", wildcard="PNG 이미지 (*.png)|*.png", style=wx.FD_SAVE|wx.FD_OVERWRITE_PROMPT)
        res = dlg.ShowModal()
        filepath = dlg.GetPath()
        dlg.Destroy()
        if res != wx.ID_OK:
            return
        self.fig.savefig(filepath, dpi=300)
        wx.MessageBox("이미지를 저장하였습니다.", "안내", wx.OK|wx.ICON_INFORMATION)

    # ----------------------------------------------------------------------
    # 차트 드로잉 및 크기 조정
    # ----------------------------------------------------------------------

    def draw(self):
        """현재 UI 상태에 따라 적절한 차트 드로잉 메서드를 호출합니다."""
        self.ax.clear()

        if self.__rb_both.GetValue():
            self.__draw_both()
        elif self.__rb_actual.GetValue():
            self.__draw_single(True) # 실적
        else:
            self.__draw_single(False) # 계획
            
        self.__update_canvas_size()
        self.__canvas.draw()

    def __update_canvas_size(self):
        """
        데이터 개수에 따라 높이(픽셀)만 계산해서 Figure에 할당합니다.
        좌우 여백은 'constrained layout'이 알아서 처리합니다.
        """
        row_count = len(self.__data_to_draw)
        
        if row_count == 0:
            total_height_px = 200
        else:
            # 여백 계산을 단순화 (상하 패딩 정도만 추가)
            total_height_px = (row_count * self.BAR_HEIGHT_PX) + self.VERTICAL_PADDING_PX

        # 스크롤 윈도우의 현재 너비
        width_px = self.__scroll.GetClientSize().width
        if width_px < 50: width_px = 600

        dpi = self.fig.get_dpi()
        
        # Figure 사이즈 적용 (inch)
        self.fig.set_size_inches(width_px / dpi, total_height_px / dpi, forward=True)
        
        # Canvas 위젯 MinSize 적용
        self.__canvas.SetMinSize((width_px, int(total_height_px)))
        self.__scroll.FitInside()

    def __get_filtered_sorted_data(self, mode: Literal["모두", "계획", "실적"]) -> dict[str, tuple[float, float] | float]:
        data = {}
        for item in self.__items:
            # 항목명 추출
            label = item.category.name if isinstance(item, ItemCategory) else item.ctr.name
            
            # 체크되지 않은 항목은 제외
            if not self.__cks[label].GetValue():
                continue
            
            match mode:
                case "모두":
                    data[label] = (item.plan, item.actual)
                case "계획":
                    data[label] = item.plan
                case "실적":
                    data[label] = item.actual

        # 합계 옵션 처리
        if data:
            if mode == "모두":
                if self.__ck_sum.GetValue():
                    # (계획, 실적)의 합계
                    plans = [v[0] for v in data.values()]
                    actuals = [v[1] for v in data.values()]
                    s = (np.nansum(plans), np.nansum(actuals))
                    data = {"합계": s, **data}
                # 정렬 (계획 기준 오름차순) -> barh에서는 오름차순이 bottom부터 그려짐
                data = dict(sorted(data.items(), key=lambda kv: kv[1][0]))
            else:
                if self.__ck_sum.GetValue():
                    # 단일 값의 합계
                    data["합계"] = np.nansum(list(data.values()))
                # 정렬 (값 기준 오름차순)
                data = dict(sorted(data.items(), key=lambda kv: kv[1]))
                
        # 리팩토링된 draw()에서 사용할 수 있도록 데이터를 저장
        self.__data_to_draw = data
        return data

    def __draw_single(self, actual: bool):
        data = self.__get_filtered_sorted_data("실적" if actual else "계획")
        
        if not data:
            self.ax.set_yticks([])
            self.ax.set_xticks([])
            for spine in self.ax.spines.values():
                spine.set_visible(False)
            self.ax.text(0.5, 0.5, "데이터 없음", ha="center", va="center")
            return
            
        labels = list(data.keys())
        values = list(data.values())
        y = np.arange(len(labels))

        self.ax.barh(y, values, height=0.6, color=COLORMAP[0], edgecolor="none")
        
        self.ax.set_yticks(y)
        self.ax.set_yticklabels(labels)
        
        # 스타일 적용 (Bold)
        self.ax.tick_params(axis='y', labelsize=10, left=False) # left=False로 눈금(tick) 선은 숨김
        # for label in self.ax.get_yticklabels():
        #     label.set_fontweight('bold')

        # 값 텍스트 표시
        max_v = max(values) if values else 1
        for yi, val in enumerate(values):
            self.ax.text(val + (max_v * 0.01), yi, simplify_won(val), ha='left', va='center')

        self.__finalize_axes(len(labels), max_v)

    def __draw_both(self):
        data = self.__get_filtered_sorted_data("모두")
        
        if not data:
            self.ax.set_yticks([])
            self.ax.set_xticks([])
            for spine in self.ax.spines.values():
                spine.set_visible(False)
            self.ax.text(0.5, 0.5, "데이터 없음", ha="center", va="center")
            return

        labels = list(data.keys())
        plans  = [v[0] for v in data.values()]
        actual = [v[1] for v in data.values()]
        
        y = np.arange(len(labels))

        self.ax.barh(y, plans,  height=0.6, color=COLORMAP[0], label='계획')
        self.ax.barh(y, actual, height=0.3, color=COLORMAP[2], label='실적')

        # [핵심 변경 2] ax.text 대신 yticks 사용
        self.ax.set_yticks(y)
        self.ax.set_yticklabels(labels)
        
        self.ax.tick_params(axis='y', labelsize=10, left=False)
        # for label in self.ax.get_yticklabels():
        #     label.set_fontweight('bold')

        max_val = max(plans + actual) if plans + actual else 1
        
        for yi, (p, a) in enumerate(zip(plans, actual)):
            xx = max(p, a) + (max_val * 0.02)
            text = f"{simplify_won(a)} / {simplify_won(p)}\n"
            if p > 0:
                text += f"{a/p*100:0.1f}%"
            else:
                text += "-"
            self.ax.text(xx, yi, text, ha='left', va='center', fontsize=9)

        self.__finalize_axes(len(labels), max_val)

    def __finalize_axes(self, item_count, max_val):
        """축의 공통 설정 (테두리 제거, 범위 설정)"""
        self.ax.set_xticks([])
        
        for spine in self.ax.spines.values():
            spine.set_visible(False)
            
        self.ax.set_ylim(-0.5, item_count - 0.5)
        self.ax.set_xlim(0, max_val * 1.3)

class PanelViewer(wx.Panel):
    def __init__(self, parent: wx.Panel):
        wx.Panel.__init__(self, parent)
        self.__selected_category_node = None
        self.__selected_ctr_node = None
        self.__expanded_category = True
        self.__expanded_ctr = True
        self.__ctr_filter: CostCtr|None = None
        self.__category_filter: CostCategory|None = None
        self.__set_layout()
        self.__bind_events()
        self.redraw_trees()
        self.update_values()

    def __set_layout(self):
        pn_top = wx.Panel(self)

        periods = ["전체", "1Q", "2Q", "3Q", "4Q"]
        periods.extend([f"{i}월" for i in range(1, 13)])
        cb_period = wx.ComboBox(pn_top, value="전체", choices=periods, style=wx.CB_READONLY)

        units = ["자동", "억원", "백만원", "천원", "원"]
        cb_unit = wx.ComboBox(pn_top, value="자동", choices=units, style=wx.CB_READONLY)

        sz_horz = wx.BoxSizer(wx.HORIZONTAL)
        sz_horz.AddMany((
            (cb_period, 0, wx.ALIGN_CENTER_VERTICAL), ((10, -1), 0),
            (cb_unit, 0, wx.ALIGN_CENTER_VERTICAL)
        ))
        sz_top = wx.BoxSizer(wx.HORIZONTAL)
        sz_top.Add(sz_horz, 0)
        pn_top.SetSizer(sz_top)

        pn_bot = ScrolledPanel(self)
        pn_bot.SetupScrolling(True, False)
        tc_ctr_filter = wx.TextCtrl(pn_bot, size=(300, -1), style=wx.TE_READONLY)
        bt_category_up = wx.Button(pn_bot, label="▲", style=wx.BU_EXACTFIT)
        bt_category_down = wx.Button(pn_bot, label="▼", style=wx.BU_EXACTFIT)
        bt_set_category_filter = wx.Button(pn_bot, label="선택한 Category로 필터")
        bt_expand_category = wx.Button(pn_bot, label="펼치기/접기")
        sz_category_header = wx.BoxSizer(wx.HORIZONTAL)
        sz_category_header.AddMany((
            (tc_ctr_filter, 0, wx.ALIGN_CENTER_VERTICAL), ((20, -1), 1),
            (bt_category_up, 0, wx.ALIGN_CENTER_VERTICAL), ((3, -1), 0),
            (bt_category_down, 0, wx.ALIGN_CENTER_VERTICAL), ((3, -1), 0),
            (bt_set_category_filter, 0, wx.ALIGN_CENTER_VERTICAL), ((3, -1), 0),
            (bt_expand_category, 0, wx.ALIGN_CENTER_VERTICAL)
        ))
        tr_category = TreeCategory(pn_bot)
        sz_category = wx.BoxSizer(wx.VERTICAL)
        sz_category.AddMany((
            (sz_category_header, 0, wx.EXPAND), ((-1, 3), 0),
            (tr_category, 1)
        ))

        tc_category_filter = wx.TextCtrl(pn_bot, size=(300, -1), style=wx.TE_READONLY)
        bt_ctr_up = wx.Button(pn_bot, label="▲", style=wx.BU_EXACTFIT)
        bt_ctr_down = wx.Button(pn_bot, label="▼", style=wx.BU_EXACTFIT)
        bt_set_ctr_filter = wx.Button(pn_bot, label="선택한 Ctr로 필터")
        bt_expand_ctr = wx.Button(pn_bot, label="펼치기/접기")

        target = pn_bot.FromDIP(14)
        bt_chat_gpt = wx.Button(pn_bot, label="ChatGPT 분석")
        bundle = wx.BitmapBundle.FromSVG(OPENAI_MARK_SVG.encode("utf-8"), wx.Size(target, target))
        bt_chat_gpt.SetBitmap(bundle)
        bt_chat_gpt.SetBitmapPosition(wx.LEFT)
        bt_chat_gpt.SetBitmapMargins(pn_bot.FromDIP(8), 0)

        bt_claude = wx.Button(pn_bot, label="Claude 분석")
        bundle = wx.BitmapBundle.FromSVG(CLAUDE_MARK_SVG.encode("utf-8"), wx.Size(target, target))
        bt_claude.SetBitmap(bundle)
        bt_claude.SetBitmapPosition(wx.LEFT)
        bt_claude.SetBitmapMargins(pn_bot.FromDIP(8), 0)

        sz_ctr_header = wx.BoxSizer(wx.HORIZONTAL)
        sz_ctr_header.AddMany((
            (tc_category_filter, 0, wx.ALIGN_CENTER_VERTICAL), ((20, -1), 1),
            (bt_chat_gpt, 0, wx.ALIGN_CENTER_VERTICAL), ((3, -1), 0),
            (bt_claude, 0, wx.ALIGN_CENTER_VERTICAL), ((3, -1), 0),
            (bt_ctr_up, 0, wx.ALIGN_CENTER_VERTICAL), ((3, -1), 0),
            (bt_ctr_down, 0, wx.ALIGN_CENTER_VERTICAL), ((3, -1), 0),
            (bt_set_ctr_filter, 0, wx.ALIGN_CENTER_VERTICAL), ((3, -1), 0),
            (bt_expand_ctr, 0, wx.ALIGN_CENTER_VERTICAL)
        ))
        tr_ctr = TreeCtr(pn_bot)
        sz_ctr = wx.BoxSizer(wx.VERTICAL)
        sz_ctr.AddMany((
            (sz_ctr_header, 0, wx.EXPAND), ((-1, 3), 0),
            (tr_ctr, 1)
        ))

        sz_bot = wx.BoxSizer(wx.HORIZONTAL)
        sz_bot.AddMany((
            (sz_category, 0, wx.EXPAND), ((10, -1), 0),
            (sz_ctr, 0, wx.EXPAND)
        ))
        pn_bot.SetSizer(sz_bot)

        sz_vert = wx.BoxSizer(wx.VERTICAL)
        sz_vert.AddMany((
            (pn_top, 0, wx.EXPAND), ((-1, 5), 0),
            (pn_bot, 1, wx.EXPAND)
        ))
        sz = wx.BoxSizer(wx.HORIZONTAL)
        sz.Add(sz_vert, 1, wx.EXPAND|wx.ALL, 10)
        self.SetSizer(sz)

        self.__cb_period   = cb_period  
        self.__cb_unit     = cb_unit    
        self.__tr_category = tr_category
        self.__tr_ctr      = tr_ctr     
        self.__tc_ctr_filter          = tc_ctr_filter
        self.__bt_chat_gpt            = bt_chat_gpt
        self.__bt_claude              = bt_claude  
        self.__bt_category_up         = bt_category_up  
        self.__bt_category_down       = bt_category_down
        self.__bt_set_category_filter = bt_set_category_filter
        self.__bt_expand_category     = bt_expand_category    
        self.__tc_category_filter     = tc_category_filter    
        self.__bt_ctr_up              = bt_ctr_up  
        self.__bt_ctr_down            = bt_ctr_down
        self.__bt_set_ctr_filter      = bt_set_ctr_filter
        self.__bt_expand_ctr          = bt_expand_ctr

        self.__popup_menu_category = wx.Menu()
        self.__menu_category_chart = self.__popup_menu_category.Append(wx.ID_ANY, "차트 보기")
        self.__menu_category_excel = self.__popup_menu_category.Append(wx.ID_ANY, "엑셀로 저장")

        self.__popup_menu_ctr = wx.Menu()
        self.__menu_ctr_chart = self.__popup_menu_ctr.Append(wx.ID_ANY, "차트 보기")
        self.__menu_ctr_excel = self.__popup_menu_ctr.Append(wx.ID_ANY, "엑셀로 저장")

    def __bind_events(self):
        self.__cb_period.Bind(wx.EVT_COMBOBOX, self.__on_period)
        self.__cb_unit  .Bind(wx.EVT_COMBOBOX, self.__on_unit  )
        self.__tr_category.Bind(DV.EVT_DATAVIEW_ITEM_CONTEXT_MENU, self.__on_right_click_category)
        self.__tr_ctr.Bind(DV.EVT_DATAVIEW_ITEM_CONTEXT_MENU, self.__on_right_click_ctr)
        self.__bt_chat_gpt.Bind(wx.EVT_BUTTON, self.__on_chat_gpt)
        self.__bt_claude  .Bind(wx.EVT_BUTTON, self.__on_claude  )
        self.__bt_category_up    .Bind(wx.EVT_BUTTON, self.__on_category_up  )
        self.__bt_category_down  .Bind(wx.EVT_BUTTON, self.__on_category_down)
        self.__bt_expand_category.Bind(wx.EVT_BUTTON, self.__on_expand_category)
        self.__bt_expand_ctr     .Bind(wx.EVT_BUTTON, self.__on_expand_ctr)
        self.__bt_ctr_up         .Bind(wx.EVT_BUTTON, self.__on_ctr_up  )
        self.__bt_ctr_down       .Bind(wx.EVT_BUTTON, self.__on_ctr_down)
        self.__bt_set_category_filter.Bind(wx.EVT_BUTTON, self.__on_set_category_filter)
        self.__bt_set_ctr_filter     .Bind(wx.EVT_BUTTON, self.__on_set_ctr_filter)
        self.Bind(wx.EVT_MENU, self.__on_menu_category_chart, self.__menu_category_chart)
        self.Bind(wx.EVT_MENU, self.__on_menu_category_excel, self.__menu_category_excel)
        self.Bind(wx.EVT_MENU, self.__on_menu_ctr_chart, self.__menu_ctr_chart)
        self.Bind(wx.EVT_MENU, self.__on_menu_ctr_excel, self.__menu_ctr_excel)

    def __on_period(self, event):
        _Config.PERIOD = self.__cb_period.GetValue()
        self.update_values()

    def __on_unit(self, event):
        _Config.UNIT = self.__cb_unit.GetValue()
        self.Refresh()

    def __on_right_click_category(self, event):
        item = event.GetItem()
        if not item.IsOk():
            self.__selected_category_node = None
            return
        tree = self.__tr_category
        node_id = int(item.GetID())
        node = tree.model.nodes[node_id]
        if not node.children:
            self.__selected_category_node = None
            return
        self.__selected_category_node = node
        tree.PopupMenu(self.__popup_menu_category)

    def __on_menu_category_chart(self, event):
        if not self.__selected_category_node:
            return
        items = [node.item for node in self.__selected_category_node.children]
        dlg = DialogChart(self, self.__selected_category_node.item.category.name, items)
        dlg.ShowModal()
        dlg.Destroy()

    def __on_menu_category_excel(self, event):
        if not self.__selected_category_node:
            wx.MessageBox("노드를 선택하세요.", "안내", parent=self)
            return
        self._on_save_excel(self.__tr_category, self.__selected_category_node)

    def __on_right_click_ctr(self, event):
        item = event.GetItem()
        if not item.IsOk():
            self.__selected_ctr_node = None
            return
        tree = self.__tr_ctr
        node_id = int(item.GetID())
        node = tree.model.nodes[node_id]
        if not node.children:
            self.__selected_ctr_node = None
            return
        self.__selected_ctr_node = node
        tree.PopupMenu(self.__popup_menu_ctr)

    def __on_menu_ctr_chart(self, event):
        if not self.__selected_ctr_node:
            return
        items = [node.item for node in self.__selected_ctr_node.children]
        dlg = DialogChart(self, self.__selected_ctr_node.item.ctr.name, items)
        dlg.ShowModal()
        dlg.Destroy()

    def __on_menu_ctr_excel(self, event):
        if not self.__selected_ctr_node:
            wx.MessageBox("노드를 선택하세요.", "안내", parent=self)
            return
        self._on_save_excel(self.__tr_ctr, self.__selected_ctr_node)

    def _on_save_excel(self, tree: TreeListCtrl, node: TreeListNode):
        dlg = wx.FileDialog(self, "엑셀 저장", wildcard="엑셀 파일 (*.xlsx)|*.xlsx", style=wx.FD_SAVE|wx.FD_OVERWRITE_PROMPT)
        ret = dlg.ShowModal()
        filepath = dlg.GetPath()
        dlg.Destroy()
        selected_item = tree.model.get_view_item(node)
        if ret != wx.ID_OK \
            or not selected_item.IsOk():
            return
        
        def get_value_of_node_category(n: TreeListNode, column: int):
            item: ItemCategory = n.item
            match column:
                case 0:
                    return item.category.name
                case 1:
                    return item.plan
                case 2:
                    return item.actual
                case 3:
                    return item.rem
                case 4:
                    return item.exe

        def get_value_of_node_ctr(n: TreeListNode, column: int):
            item: ItemCtr = n.item
            match column:
                case 0:
                    return item.ctr.name
                case 1:
                    return "" if item.total else item.ctr.code
                case 2:
                    return "" if item.total else item.ctr.rnd
                case 3:
                    return "" if item.total else item.ctr.oe
                case 4:
                    return item.plan
                case 5:
                    return item.actual
                case 6:
                    return item.rem
                case 7:
                    return item.exe
        
        get_method = get_value_of_node_category if isinstance(node.item, ItemCategory) else get_value_of_node_ctr

        result = []
        def work():
            try:
                wb = xl.Workbook()
                ws = wb.worksheets[0]

                # ---- (1) 헤더 작성 ----
                col_count = tree.GetColumnCount()
                headers = [tree.GetColumn(i).GetTitle() for i in range(col_count)]
                ws.append(headers)

                # --- (2) 트리 재귀 순회 함수 정의 ---
                def traverse(item: DV.DataViewItem, depth: int):
                    """
                    node: 현재 DataViewItem
                    depth: 트리 깊이 (시작 item을 0으로)
                    """
                    node_id = int(item.GetID())
                    node = tree.model.nodes[node_id]

                    # (2-1) 현재 노드 한 줄 쓰기
                    row = []
                    for col in range(col_count):
                        # 첫 번째 컬럼에만 indent 적용
                        if col == 0:
                            indent_str = " " * (depth * 2)   # depth당 4칸 공백
                            row.append(f"{indent_str}{get_method(node, col)}")
                        else:
                            row.append(get_method(node, col))

                    ws.append(row)

                    # (2-2) 자식 노드 순회
                    # PyDataViewModel: GetChildren(self, item, children)
                    # DataViewTreeStore: GetChildren(self, item) -> list
                    children = []
                    count = tree.model.GetChildren(item, children)  # children 리스트를 채워주는 형태
                    iterable = children[:count]

                    for child in iterable:
                        if isinstance(child, DV.DataViewItem):
                            traverse(child, depth + 1)

                # --- (3) 시작 item 이하 전체 순회 ---
                traverse(selected_item, depth=0)

                # --- (A) column width 설정 (DataViewCtrl의 실제 폭 기준 비례 적용) ---
                # DataViewCtrl column width는 px 단위, Excel은 약 1 = 7px 정도라고 보면 됨
                for i in range(col_count):
                    dv_width = tree.GetColumn(i).GetWidth()    # pixel
                    excel_width = dv_width * 0.15             # 적절한 비율 (조정 가능)
                    col_letter = get_column_letter(i + 1)
                    ws.column_dimensions[col_letter].width = excel_width

                # --- (B) 중앙 정렬 및 테두리 ---
                thin = Side(border_style="thin", color="000000")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                center = Alignment(horizontal="center", vertical="center")

                max_row = ws.max_row
                max_col = ws.max_column

                for r in range(1, max_row + 1):
                    for c in range(1, max_col + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.border = border
                        if r == 1 or c != 1:
                            cell.alignment = center

                # --- (C) 헤더 배경색 ---
                header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                for c in range(1, max_col + 1):
                    ws.cell(row=1, column=c).fill = header_fill

                wb.save(filepath)
                wb.close()

            except Exception as err:
                result.append(err)

            else:
                result.append(None)
        
        dlg = wx.ProgressDialog("안내", "엑셀 파일을 생성 중입니다.", parent=self)
        dlg.Pulse()
        Thread(target=work, daemon=True).start()
        while not result:
            wx.YieldIfNeeded()
            time.sleep(0.01)
        ret = result.pop(0)
        dlg.Destroy()

        def done():
            if ret is None:
                wx.MessageBox("엑셀 저장을 완료했습니다.", "안내", parent=self)
            else:
                wx.MessageBox(f"엑셀 저장 도중 오류가 발생했습니다.\n{ret}", "안내", parent=self)
            tree.Refresh()
        wx.CallAfter(done)

    def __on_category_up(self, event):
        tr = self.__tr_category
        node = tr.get_selected_node()
        if not node:
            return
        tr.move_node(node, False)

    def __on_category_down(self, event):
        tr = self.__tr_category
        node = tr.get_selected_node()
        if not node:
            return
        tr.move_node(node, True)

    def __on_ctr_up(self, event):
        tr = self.__tr_ctr
        node = tr.get_selected_node()
        if not node:
            return
        tr.move_node(node, False)

    def __on_ctr_down(self, event):
        tr = self.__tr_ctr
        node = tr.get_selected_node()
        if not node:
            return
        tr.move_node(node, True)

    def __on_expand_category(self, event):
        self.__expanded_category = not self.__expanded_category
        tr = self.__tr_category
        flag = self.__expanded_category
        logical_root = tr.get_logical_root()
        if not flag:
            for root in logical_root.children:
                tr.expand_node(root, True)
                for child in root.children:
                    tr.expand_node(child, False)
        else:
            for node in tr.model.nodes.values():
                tr.expand_node(node, True)

    def __on_expand_ctr(self, event):
        self.__expanded_ctr = not self.__expanded_ctr
        tr = self.__tr_ctr
        flag = self.__expanded_ctr
        logical_root = tr.get_logical_root()
        if not flag:
            for root in logical_root.children:
                tr.expand_node(root, True)
                for child in root.children:
                    tr.expand_node(child, False)
        else:
            for node in tr.model.nodes.values():
                tr.expand_node(node, True)

    def __on_set_category_filter(self, event):
        node = self.__tr_category.get_selected_node()
        if not node:
            wx.MessageBox("필터링할 Cost Category를 선택하세요.", "안내")
            return
        if self.__category_filter and node.item.category.pk == self.__category_filter.pk:
            return
        self.set_category_filter(CostCategory.get(node.item.category.pk))

    def __on_set_ctr_filter(self, event):
        node = self.__tr_ctr.get_selected_node()
        if not node:
            wx.MessageBox("필터링할 Cost Ctr을 선택하세요.", "안내")
            return
        if self.__ctr_filter and node.item.ctr.code == self.__ctr_filter.code:
            return
        self.set_ctr_filter(CostCtr.get(node.item.ctr.code))

    def __on_chat_gpt(self, event):
        self.__on_ai("ChatGPT")
    
    def __on_claude(self, event):
        self.__on_ai("Claude")

    def __on_ai(self, ai_type: Literal["ChatGPT", "Claude"]):
        if LoadedData.df.empty:
            wx.MessageBox("분석할 데이터가 없습니다.\n먼저 데이터를 로드하세요.", "안내", parent=self)
            return
        node = self.__tr_ctr.get_selected_node()
        if not node:
            wx.MessageBox("분석할 Cost Ctr을 선택하세요.", "안내", parent=self)
            return
        ctr = CostCtr.get(node.item.ctr.code)

        match ai_type:
            case "ChatGPT":
                key = Config.OPENAI_API_KEY
                if not key:
                    wx.MessageBox("OpenAI API Key를 설정하세요.", "안내", parent=self)
                    return
                dlg = DialogModels(self, Config.GPT_MODELS, Config.LAST_USED_GPT_MODEL)
                ret = dlg.ShowModal()
                model = dlg.get_model()
                dlg.Destroy()
                if ret != wx.ID_OK:
                    return
                Config.LAST_USED_GPT_MODEL = model
                analyze = analyze_by_gpt
            case "Claude":
                key = Config.CLAUDE_API_KEY
                if not key:
                    wx.MessageBox("Claude API Key를 설정하세요.", "안내")
                    return
                dlg = DialogModels(self, Config.CLAUDE_MODELS, Config.LAST_USED_CLAUDE_MODEL)
                ret = dlg.ShowModal()
                model = dlg.get_model()
                dlg.Destroy()
                if ret != wx.ID_OK:
                    return
                Config.LAST_USED_CLAUDE_MODEL = model
                analyze = analyze_by_claude
            case _:
                raise RuntimeError
        
        dlgp = wx.ProgressDialog("안내", f"{ai_type}을 이용하여 분석 중입니다.", parent=self)
        dlgp.Pulse()

        def on_success(json_data: str, result: str):
            dlgp.Destroy()
            wx.Yield()
            dlg = DialogAIResult(self, ai_type, model, json_data, result)
            wx.MessageBox("분석을 완료했습니다.", "안내", parent=self)
            wx.Yield()
            dlg.ShowModal()
            dlg.Destroy()
        
        def on_fail(msg: str):
            dlgp.Destroy()
            wx.Yield()
            wx.MessageBox(msg, "안내", parent=self)

        def work():
            try:
                ctrs = ctr.get_descendant()
                ctr_codes = set([c.code for c in ctrs])

                cat = self.__category_filter or CostCategory.get_root_category(False)
                category_descendant = cat.get_descendant() # type: ignore
                elements = CostElement.get_involved_in_categories(category_descendant)
                element_codes = set([elem.code for elem in elements])
                # node = self.__tr_category.get_selected_node()
                # if node:
                #     cat = CostCategory.get(node.item.category.pk)
                #     elements = CostElement.get_involved_in_categories(cat.get_descendant())
                # else:
                #     elements = CostElement.get_involved_in_categories(list(CostCategory.get_all().values()))
                # element_codes = set([elem.code for elem in elements])

                filtered_df = LoadedData.df.loc[
                    LoadedData.df["Cost Center"].isin(ctr_codes) \
                    & LoadedData.df["Cost Element"].isin(element_codes)
                ]

                assert not filtered_df.empty, "분석할 데이터가 없습니다."

                months = Config.get_months()
                # all_categories = CostCategory.get_all()
                all_elements = CostElement.get_all()
                all_ctrs = CostCtr.get_all()

                category_list = [
                    _CostCategory(
                        pk=cat.pk,
                        parent_pk=cat.parent_pk,
                        name=cat.name
                    ) for cat in category_descendant
                ]
                element_list = [
                    _CostElement(
                        code=elem.code,
                        description=elem.description,
                        category_pk=elem.category_pk
                    ) for elem in all_elements.values() if elem.code in element_codes
                ]
                ctr_list = [
                    _CostCtr(
                        code=ctr.code,
                        parent_code=ctr.parent_code,
                        name=ctr.name,
                        rnd=["Research", "Develop"].index(ctr.rnd),
                        oe=["공통비", "RE", "OE"].index(ctr.oe)
                    ) for ctr in all_ctrs.values()
                ]

                df = filtered_df.fillna(0)
                cols = {name: i for i, name in enumerate(df.columns)}
                by_elem_code = {}
                by_ctr_code = {}
                for row in df.itertuples(index=False, name=None):
                    ctr_code = row[cols["Cost Center"]]
                    elem_code = row[cols["Cost Element"]]
                    if ctr_code not in ctr_codes \
                        or elem_code not in all_elements:
                        continue
                    if elem_code not in by_elem_code:
                        by_elem_code[elem_code] = {"code": elem_code, "planned": 0, "executed": 0}
                    if ctr_code not in by_ctr_code:
                        by_ctr_code[ctr_code] = {"code": ctr_code, "planned": 0, "executed": 0}
                    for month in months:
                        planned = row[cols[f"ConvPlan({month})"]]
                        executed = row[cols[f"ConvActual({month})"]]
                        by_elem_code[elem_code]["planned"] += planned
                        by_elem_code[elem_code]["executed"] += executed
                        by_ctr_code[ctr_code]["planned"] += planned
                        by_ctr_code[ctr_code]["executed"] += executed

                system_prompt, user_prompt, json_data = get_prompts_for_ai(
                    category_list,
                    element_list,
                    ctr_list,
                    [
                        _BudgetByElement(
                            cost_element_code=item["code"],
                            planned=item["planned"],
                            executed=item["executed"]
                        ) for item in by_elem_code.values()
                    ],
                    [
                        _BudgetByCtr(
                            cost_ctr_code=item["code"],
                            planned=item["planned"],
                            executed=item["executed"]
                        ) for item in by_ctr_code.values()
                    ]
                )

                res = analyze(system_prompt, user_prompt, key, model)

            except Exception as err:
                if isinstance(err, AssertionError):
                    msg = str(err)
                else:
                    msg = f"AI 분석 중 오류가 발생했습니니다.\n\n{format_exc()}"
                wx.CallAfter(on_fail, msg)
            
            else:
                wx.CallAfter(on_success, json_data, res)

        Thread(target=work, daemon=True).start()

    def redraw_trees(self):
        """노드를 초기화 후 재생성"""
        tr = self.__tr_category
        tr.clear_nodes()
        for cat in LoadedData.cached_cost_category.values():
            tr.add_node(
                tr.get_node_by_key(cat.parent_pk) if cat.parent_pk is not None else None,
                cat.pk,
                ItemCategory(cat)
            )

        tr = self.__tr_ctr
        tr.clear_nodes()
        for ctr in LoadedData.cached_cost_ctr.values():
            level = LoadedData.get_level_of_ctr_from_cache(ctr)
            if level == 3:
                tr.add_node(
                    ctr.parent_code and tr.get_node_by_key(f"TOTAL-{ctr.parent_code}"),
                    ctr.code,
                    ItemCtr(ctr, False)
                )
                continue
            node = tr.add_node(
                ctr.parent_code and tr.get_node_by_key(f"TOTAL-{ctr.parent_code}"),
                f"TOTAL-{ctr.code}",
                ItemCtr(ctr, True)
            )
            tr.add_node(
                node,
                ctr.code,
                ItemCtr(ctr, False)
            )

    def update_values(self):
        """노드는 유치한 체로 plan과 actual 값을 재계산"""
        if not self.__category_filter:
            self.__category_filter = CostCategory.get_root_category(False)
        if not self.__ctr_filter:
            self.__ctr_filter = CostCtr.get_root_ctr()
        self.__tc_category_filter.SetValue(" > ".join([cat.name for cat in self.__category_filter.get_path()]))
        self.__tc_ctr_filter.SetValue(" > ".join([ctr.name for ctr in self.__ctr_filter.get_path()]))
        df = LoadedData.get_filtered_df()
        category_descendant = self.__category_filter.get_descendant()
        elements = CostElement.get_involved_in_categories(category_descendant)
        element_codes = [elem.code for elem in elements]
        ctr_descendant = self.__ctr_filter.get_descendant()
        ctr_codes = [ctr.code for ctr in ctr_descendant]
        mask_element = df["Cost Element"].isin(element_codes)
        mask_ctr = df["Cost Center"].isin(ctr_codes)
        period = self.__cb_period.GetValue()
        match period:
            case "전체":
                months = list(range(1, 13))
            case "1Q":
                months = [1, 2, 3]
            case "2Q":
                months = [4, 5, 6]
            case "3Q":
                months = [7, 8, 9]
            case "4Q":
                months = [10, 11, 12]
            case _:
                months = [int(period[:-1]),]
        cols_plan = [f"ConvPlan({i})" for i in months]
        cols_actual = [f"ConvActual({i})" for i in months]

        tr = self.__tr_category
        pk_vs_amounts = {}
        for nid, node in tr.model.nodes.items():
            item: ItemCategory = node.item
            if item is None:
                continue
            category: CostCategory = item.category
            level = LoadedData.get_level_of_category_from_cache(category)
            if level < MAXIMUM_DEPTH_OF_CATEGORY:
                if category.pk not in pk_vs_amounts:
                    pk_vs_amounts[category.pk] = {"plan": 0, "actual": 0}
                continue
            element_codes_of_category = [
                elem_code 
                for elem_code, elem in LoadedData.cached_cost_element.items()
                if elem.category_pk == category.pk
            ]
            mask = (df["Cost Element"].isin(element_codes_of_category)) & (mask_ctr)
            plan = df[mask][cols_plan].fillna(0).sum().sum()
            actual = df[mask][cols_actual].fillna(0).sum().sum()
            pk_vs_amounts[category.pk] = {"plan": plan, "actual": actual}
            parent_pk = category.parent_pk
            while parent_pk is not None:
                pk_vs_amounts[parent_pk]["plan"] += plan
                pk_vs_amounts[parent_pk]["actual"] += actual
                parent_pk = LoadedData.cached_cost_category[parent_pk].parent_pk
        for nid, node in tr.model.nodes.items():
            item: ItemCategory = node.item
            if item is None:
                continue
            category = item.category
            plan   = pk_vs_amounts[category.pk]["plan"]
            actual = pk_vs_amounts[category.pk]["actual"]
            item.plan   = plan  
            item.actual = actual
            if plan is None:
                item.rem = None
                item.exe = None
            elif actual is None:
                item.rem = plan
                item.exe = None
            else:
                item.rem = plan-actual
                item.exe = None
                if plan:
                    item.exe = actual/plan
            tr.update_node(node)

        tr = self.__tr_ctr
        code_vs_amounts = {}
        for nid, node in tr.model.nodes.items():
            item: ItemCtr = node.item
            if item is None:
                continue
            ctr: CostCtr = item.ctr
            if item.total:
                code_vs_amounts[f"TOTAL-{ctr.code}"] = {"plan": 0, "actual": 0}
                continue
            mask = (df["Cost Center"] == ctr.code) & (mask_element)
            plan = int(df[mask][cols_plan].fillna(0).sum().sum())
            actual = int(df[mask][cols_actual].fillna(0).sum().sum())
            code_vs_amounts[ctr.code] = {"plan": plan, "actual": actual}
            key = f"TOTAL-{ctr.code}"
            if key in code_vs_amounts:
                code_vs_amounts[key]["plan"] += plan
                code_vs_amounts[key]["actual"] += actual
            parent_code = ctr.parent_code
            while parent_code:
                parent_ctr = LoadedData.cached_cost_ctr[parent_code]
                key = f"TOTAL-{parent_ctr.code}"
                code_vs_amounts[key]["plan"] += plan
                code_vs_amounts[key]["actual"] += actual
                parent_code = parent_ctr.parent_code
        for nid, node in tr.model.nodes.items():
            item: ItemCtr = node.item
            if item is None:
                continue
            ctr = item.ctr
            if item.total:
                key = f"TOTAL-{ctr.code}"
            else:
                key = ctr.code
            plan   = code_vs_amounts[key]["plan"]
            actual = code_vs_amounts[key]["actual"]
            item.plan   = plan  
            item.actual = actual
            if plan is None:
                item.rem = None
                item.exe = None
            elif actual is None:
                item.rem = plan
                item.exe = None
            else:
                item.rem = plan-actual
                item.exe = None
                if plan:
                    item.exe = actual/plan
            tr.update_node(node)

    def set_ctr_filter(self, ctr: CostCtr):
        self.__ctr_filter = ctr
        self.update_values()

    def set_category_filter(self, category: CostCategory):
        self.__category_filter = category
        self.update_values()


